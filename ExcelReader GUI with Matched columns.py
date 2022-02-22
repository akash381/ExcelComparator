import sys
from sys import exit
import openpyxl
from openpyxl.styles import PatternFill
# from the tkinter library
from tkinter import *

# import filedialog module
from tkinter import filedialog

# import ttk for progress bar
from tkinter import ttk

import time
from datetime import datetime
from openpyxl.drawing.text import Font

#A function to compare value with Run Data and Time
def findRunDateTimeInString(value1):
    if(str(value1).find("Run Date and Time") != -1):
        return True
    return False 

# A function to compare None and blank
def CompareNoneAndBlank(value1, value2):
    if (value1 is None and not value2 ):
        return True
    elif (value2 is None and not value1):
        return True
    return False

# A function to remove unwanted stuff from comparision
def compare(value1, value2):
    if(CompareNoneAndBlank(value1, value2) or findRunDateTimeInString(value1)):
        return True
    return False

# datetime object containing current date and time
now = datetime.now()
dt_string = now.strftime("%d-%b-%Y %HH%MM%SS")

file1 = "" #Before.xlsx" #input("Enter the first file name : ") #"BEFORE PBRER 5.0.xlsx"
file2 = ""#After.xlsx" # input("Enter the Second file name : ") #"AFTER PBRER 5.0 - Copy.xlsx"
file3 = "Compared File " + dt_string + ".xlsx"
log_file = "Log File " + dt_string + ".txt"

def compareExcel():    
    #A log file having all the differences
    f = open(log_file, "w")

    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    fill_pattern_red = PatternFill(patternType = "solid", fgColor = 'FF3333')
    fill_pattern_green = PatternFill(patternType = "solid", fgColor = 'BBFF33')
    fill_pattern_yellow = PatternFill(patternType = "solid", fgColor = 'FFFF00')
    # bold_font = Font(bold = True)

    noOfSheets1 = len(wb1.sheetnames)
    noOfSheets2 = len(wb2.sheetnames)

    f.write("This is log file containing all the differences found in the comparision.")

    if (noOfSheets1 != noOfSheets2):
        print("Number of sheets are different in both the workbook")
        f.write("\n\nNumber of sheets are different in both the workbook. \n" + 
                "noOfSheets1 : " + str(noOfSheets1) + " noOfSheets2 : " + str(noOfSheets2) +
                "\nHence closing the file comarision.")
        return "No. of sheets in first file : " + str(noOfSheets1) + " No. of sheets in second file : " + str(noOfSheets2)
        # sys.exit("noOfSheets1 : " + str(noOfSheets1) + " noOfSheets2 : " + str(noOfSheets2))

    for s in range(0, noOfSheets1): # Update this to include criteria sheet : from 0 to noOfSheets1
        mismatchFound = 0
        sh1 = wb1.worksheets[s]
        sh2 = wb2.worksheets[s]

        sheetName = wb1.sheetnames[s]
        
        row1 = sh1.max_row
        row2 = sh2.max_row
        row_max = max(row1, row2)

        f.write("\n\n----------------Starting comparision for sheet : " + sheetName + "---------------\n")
        # Compare number of rows
        if (row1 != row2):
            print("Number of rows are different in both the sheet for : " + sheetName + "\n")
            f.write("Number of rows are different in both the sheet for : " + sheetName + "\n")

        column1 = sh1.max_column
        column2 = sh2.max_column
        column_max = max(column1, column2)

        # Compare number of rows
        if (column1 != column2):
            print("Number of columns are different in both the sheet for : " + sheetName + "\n")
        
        # Putting a match column
        # sh2.insert_rows(1)
        sh2.cell(1, column_max + 1).value = 'Mismatch?'
        sh2.cell(1, column_max + 1).fill = fill_pattern_yellow
        # sh2.cell(1, column_max + 1).font = bold_font

        extra_row = False
        extra_column = False

        for r in range(2, row_max + 1):  # Change this from 1 to row1 + 1 for whole scanning
            mismathRow = 0
            #sh2.insert_cols(column2 + 2)
            #print("Sheet color : " + str(sh2.cell(r, 1).fill.start_color.index))
            
            for c in range(1, column_max + 1):
                value1 = sh1.cell(r, c).value
                value2 = sh2.cell(r, c).value

                # Coloring extra columns and rows
                if(r > row1 or r > row2 ):
                    if (extra_row == False):
                        f.write("Extra row found at Row : " + str(r) + "\n")
                    sh2.cell(r, c).fill = fill_pattern_red
                    extra_row = True
                    mismathRow += 1
                    pass
                elif (c > column1 or c > column2 ):
                    if (extra_column == False):
                        f.write("Extra column found at Column : " + str(c) + "\n")
                    sh2.cell(r, c).fill = fill_pattern_red
                    extra_column = True
                    mismathRow += 1
                    pass

                #Not comparing unwanted texts
                elif(compare(value1, value2)):
                    pass

                #Comparing both cell value
                elif(value1 == value2):
                    # print("Matched : " + str(value1))
                    # if(value1 is not None):
                    #     sh2.cell(r, c).fill = fill_pattern_green
                    pass
                else:
                    f.write("Mismatch found at row " + str(r) + " column " + str(c) + " : \n" 
                        + "\t\t Before value : " + str(value1) + "\n"
                        + "\t\t After value : " + str(value2) + "\n")
                    # print("Not Matched : " + str(value1) + " , " + str(value2))
                    sh2.cell(r, c).fill = fill_pattern_red
                    mismatchFound += 1
                    mismathRow += 1

            # This can create a row at the end of the sheet saying Yes and No for Mismatches? Also uncomment line 73
            if (mismathRow > 0):
                sh2.cell(r, column_max + 1).value = 'Yes'
            else:
                sh2.cell(r, column_max + 1).value = 'No'
        if(mismatchFound == 0):
            f.write("Everything matched in this sheet.\n")
        # else:
        #     f.write("Number of mismatches in this sheet is : " + str(mismatchFound) + "\n")

    f.write("\n\n-----------------Comparision complete!--------------------")

    wb2.save(file3)
    f.close()
    return "Done"

#---------------------------------------------------------------Starting GUI---------------------#
# file explorer window

def browseFiles1():
    filename = filedialog.askopenfilename(initialdir = "./", title = "Select a File", filetypes = (("Text files", "*.xlsx*"), ("all files", "*.*")))
    global file1
    file1 = filename
    label_file_explorer.configure(text=filename)
    
def browseFiles2():
    filename = filedialog.askopenfilename(initialdir = "./", title = "Select a File", filetypes = (("Text files", "*.xlsx*"), ("all files", "*.*")))
    global file2
    file2 = filename
    label_file_explorer2.configure(text=filename)

# def before_comparing():
# 	# Change label contents
#     label_before_comparision.configure(text = 'Comparing...')
#     window.update_idletasks()
#     time.sleep(1)

def comparing():
	# Change label contents
    if(file1 == "" or file2 == ""):
        label_after_comparision.configure(text = 'Please select both the files before comparing.')
    else:
        label_after_comparision.configure(text = 'Comparing...')
        window.update_idletasks()
        output = compareExcel()
        if output == "Done":
            result = 'Comparision Complete!\n' + 'Please look Compared File for output and Log File for logs'
        else:
            result = 'Comparision cannot be done due to different number of sheet!\n' + output
        label_after_comparision.configure(text = result)
																						
# Create the root window
window = Tk()

# Set window title
window.title('Excel Comparing Software')

# Set window size
window.geometry("800x500")

#Set window background color
window.config(background = "white")

# Create a File Explorer label
label_file_explorer = Label(window,
							#text = "File Explorer using Tkinter",
							#width = 50, height = 5,
							fg = "blue")

# Create a File Explorer label
label_file_explorer2 = Label(window,
							#text = "File Explorer using Tkinter",
							#width = 50, height = 5,
							fg = "blue")

# # Create a label before comparing
# label_before_comparision = Label(window,
# 							#text = "Comparing...",
# 							#width = 50, height = 5,
# 							fg = "blue")

# Create a label after comparing
label_after_comparision = Label(window,
							# text = "Comparing...",
							#width = 50, height = 5,
							fg = "blue")   

	
button_explore = Button(window,
						text = "Browse Before File",
                        width = 15, height = 1,
						command = browseFiles1)

button_explore2 = Button(window,
						text = "Browse After File",
                        width = 15, height = 1,
						command = browseFiles2)

button_compare = Button(window,
					text = "Compare",
                    width = 15, height = 2,
					command = comparing)

button_exit = Button(window,
					text = "Exit",
                    width = 15, height = 1,
					command = exit)

# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns

button_explore.grid(column = 1, row = 1)
label_file_explorer.grid(column = 2, row = 1)

button_explore2.grid(column = 1, row = 2)
label_file_explorer2.grid(column = 2, row = 2)

button_compare.grid(column = 1, row = 3)
# label_before_comparision.grid(column = 2, row = 3)

label_after_comparision.grid(column = 2, row = 3)

button_exit.grid(column = 1, row = 4)

# Let the window wait for any events
window.mainloop()