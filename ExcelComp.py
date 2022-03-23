# System imports
import sys
from sys import exit

# Library to maipulate excel
import openpyxl
from openpyxl.styles import PatternFill

# Files to maipuate
file1 = "Test Before.xlsx"
file2 = "Test After.xlsx"

# Ouput file names
file3 = "Compared File.xlsx"
logFile = "Log File.txt"

# Log file open
f = open(logFile, "w")

# Load excels
wb1 = openpyxl.load_workbook(file1)
wb2 = openpyxl.load_workbook(file2)

# Pattern red
fill_pattern_red = PatternFill(patternType = "solid", fgColor = 'FF3333')

# Log file start
f.write("This is log file containing all the differences found in the comparision.\n")
f.write("First File Path : " + file1 + "\n")
f.write("Second File Path : " + file2 + "\n")

# number of sheets in each excel
noOfSheets1 = len(wb1.sheetnames)
noOfSheets2 = len(wb2.sheetnames)

# Compare number of sheets
if (noOfSheets1 != noOfSheets2):
    print("Number of sheets are different in both the workbook")
    f.write("\n\nNumber of sheets are different in both the workbook. \n" + 
            "noOfSheets1 : " + noOfSheets1 + " noOfSheets2 : " + noOfSheets2 +
            "\nHence closing the file comarision.")
    sys.exit("noOfSheets1 : " + str(noOfSheets1) + " noOfSheets2 : " + str(noOfSheets2))

# If number of sheet is equal in both
for s in range(0, noOfSheets1):
    mismatchFound = 0
    sh1 = wb1.worksheets[s]
    sh2 = wb2.worksheets[s]

    sheetName = wb1.sheetnames[s]

    row1 = sh1.max_row
    row2 = sh2.max_row
    row_max = max(row1, row2)

    f.write("\n\n----------------Starting comparision for sheet : " + sheetName + "---------------\n")
    # Compare number of rows
    if (row1 != row2):
        print("Number of rows are different in both the sheet for : " + sheetName + "\n")
        f.write("Number of rows are different in both the sheet for : " + sheetName + "\n")

    column1 = sh1.max_column
    column2 = sh2.max_column
    column_max = max(column1, column2)

    # Compare number of rows
    if (column1 != column2):
        print("Number of columns are different in both the sheet for : " + sheetName + "\n")
        f.write("Number of columns are different in both the sheet for :" + sheetName + "\n")

    extra_row = False
    extra_column = False

    for r in range(1, row_max + 1):
        for c in range(1, column_max + 1):
            value1 = sh1.cell(r, c).value
            value2 = sh2.cell(r, c).value
            # Coloring extra columns and rows
            if(r > row1 or r > row2 ):
                if (extra_row == False):
                    f.write("Extra row found at Row : " + str(r) + "\n")
                sh2.cell(r, c).fill = fill_pattern_red
                extra_row = True
                pass

            elif (c > column1 or c > column2 ):
                if (extra_column == False):
                    f.write("Extra column found at Column : " + str(c) + "\n")
                sh2.cell(r, c).fill = fill_pattern_red
                extra_column = True
                pass
            #Comparing both cells value
            elif(value1 == value2):
                pass

            else:
                f.write("Mismatch found at row " + str(r) + " column " + str(c) + " : \n" 
                    + "\t\t Before value : " + str(value1) + "\n"
                    + "\t\t After value : " + str(value2) + "\n")

                sh2.cell(r, c).fill = fill_pattern_red
                mismatchFound += 1
    if(mismatchFound == 0):
        f.write("Everything matched in this sheet.\n")

f.write("\n\n-----------------Comparision complete!--------------------\n\n")

wb2.save(file3)

print("Execution complete.")
print("Please check Compared File.xlsx for output")
print("And check Log File.txt for logs")

f.write("Compared File Name : " + file3 + "\n")
f.write("Log File Path : " + logFile + "\n")

f.close()
